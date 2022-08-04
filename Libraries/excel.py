import openpyxl
from openpyxl.styles import PatternFill, Border, Side
try:
    import tkinter as tk
except:
    import Tkinter as tk
from tkinter import filedialog
import re

def scrap_dictionary(ditionary):
    """Gets a dictionary file previously edited, then it graps all the information required for the Excel format"""
    #All the information grapped for the Excel file format from the dictionary is pulled as shown next:
    #   *KEY        *Order  *Nested KEY     *Value
    #    Comprobane			
    #                5       |SubTotal	     |Sub 
	#   	         6       |Total	         |Total 
	#	             8       |Serie       	 |Serie
	#	             9       |Folio	         |Folio
    #
    #    Receptor			
	#   	         1       |Rfc	         |rfc
	#	             2       |Nombre	     |Razon social
    #
    #    Concepto			
	#	             3       |ClaveProdServ	 |Clave del producto y/o servicio
	#	             4       |Descripcion	 |Concepto	
    # 	
    #    Complementos			
	#   	         7       |UUID	         |Folio fiscal
	#	             10      |FechaTimbrado	 |Fecha emision

    scrapped_file = [ditionary['Receptor']['Rfc'], ditionary['Receptor']['Nombre'],
                    ditionary['Concepto']['ClaveProdServ'], ditionary['Concepto']['Descripcion'],
                    int(float(ditionary['Comprobane']['SubTotal']) * 10), int(float(ditionary['Comprobane']['Total']) * 10),
                    ditionary['Complementos']['UUID'], ditionary['Comprobane']['Serie'], 
                    ditionary['Comprobane']['Folio'], ditionary['Complementos']['FechaTimbrado']
                    ]
    #scrapped_file[5] = int(float(scrapped_file[5])) * 10
    return scrapped_file

def get_excel_path():
    """Gets the file path from a window where file must be selected"""
    # Creates a window
    root = tk.Tk()
    root.withdraw()
    # Asks for a excel (.xlsx) file, to get its rute 
    file_path = filedialog.askopenfilename(filetypes=["Excel .xlsx"])
    print(file_path)
    # Abre el documento de Excel a revisar
    return file_path

# *NOTE* old module reused from another program
def close_path(file_path, working_sheet):
    y = ""
    for i in range(len(file_path), -1, -1):
        if file_path[i-1] == "/":
            break  # Ciclo ends when "/" is encountered.
        # Retrive the file's name from the path, but it is backward
        y += file_path[i-1]
    # 
    y = y[::-1]

    working_sheet.save("{}{}".format(file_path[0:len(file_path)-len(y)],y))

def store_invoice(file_path, file_to_append):
    """Strip all data from the the list and stored in our excel file"""

    # Create a dictionary with every month
    months_dict = {
        "-01-" : "Enero",
        "-02-" : "Febrero",
        "-03-" : "Marzo",
        "-04-" : "Abril",
        "-05-" : "Mayo",
        "-06-" : "Junio",
        "-07-" : "Julio",
        "-08-" : "Agosto",
        "-09-" : "Septiembre",
        "-10-" : "Octubre",
        "-11-" : "Noviembre",
        "-12-" : "Diciembre"
    }
    # A pattern that will be use to check for the month
    pattern = "-[0-9]{2}-"
    # From the 'date' variable, it will only look up for the pattern in 'pattern'
    get_month_pattern = re.findall(pattern, file_to_append[9])[0]
    # This method will the information from to the Excel format
    if get_month_pattern in months_dict:
        # Creates a working sheet using the file path
        working_sheet = openpyxl.load_workbook(file_path)
        # It posicionates on the document's month. Using the month pattern, get the month for the dictionary.
        sheet = working_sheet[months_dict[get_month_pattern]]
        
        print(months_dict[get_month_pattern])
        # It is used to start at 4th row, and then iterates from its position
        iterations = 4
        # It is used to break, the while loop
        exeption_file = False

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'))

        while exeption_file != True:
            # Checks if the invoice was already summited to the document
            if sheet.cell(row=iterations, column=7).value == file_to_append[6]:
                print("Already documented in {} row {}: \nID in documents: {}\nID to store: {}".format(months_dict[get_month_pattern], iterations, sheet.cell(row=iterations, column=7).value ,file_to_append[6]))
                exeption_file = True
            # If it was not registed and it is a blank space, it will insert all information from the Dictionary to the document
            elif sheet.cell(row=iterations, column=7).value is None:
                sheet.insert_rows(iterations)
                for i in range(0,10):
                    if i == 5 or i == 4:
                        sheet.cell(row=iterations, column=i+1).value = file_to_append[i] / 10
                    else:
                        sheet.cell(row=iterations, column=i+1).value = file_to_append[i]
                
                for i in range(0,12):
                    if iterations % 2 == 0:
                        sheet.cell(row=iterations, column=i+1).fill = PatternFill(fgColor="e2f0d9", fill_type = "solid")
                    sheet.cell(row=iterations, column=i+1).border = thin_border

                
                print('null')
                link_sheet = iterations
                exeption_file = True
            # Checks for the next row
            iterations += 1
        # Positioned in the sheet where all invoices are stored for analysis
        sheet_raw_invoice = working_sheet["RAW Facturas"]
        #
        exeption_file = False
        iterations = 2
        while exeption_file != True:
            # Checks if the invoice was already summited to the document
            if sheet_raw_invoice.cell(row=iterations, column=7).value == file_to_append[6]:
                print("Already documented in RAW Facturas row {}: \nID in documents: {}\nID to store: {}".format(iterations, sheet_raw_invoice.cell(row=iterations, column=7).value ,file_to_append[6]))
                exeption_file = True
            # If it was not registed and it is a blank space, it will insert all information from the Dictionary to the document
            elif sheet_raw_invoice.cell(row=iterations, column=7).value is None:
                sheet_raw_invoice.insert_rows(iterations)
                for i in range(0,10):
                    if i == 5 or i == 4:
                        sheet_raw_invoice.cell(row=iterations, column=i+1).value = file_to_append[i] / 10
                    else:
                        sheet_raw_invoice.cell(row=iterations, column=i+1).value = file_to_append[i]
                
                for i in range(0,12):
                    if iterations % 2 == 0:
                        sheet_raw_invoice.cell(row=iterations, column=i+1).fill = PatternFill(fgColor="e2f0d9", fill_type = "solid")
                    sheet_raw_invoice.cell(row=iterations, column=i+1).border = thin_border
                
                #pattern_dir = "/[\w]{n}[.]?[\w]{n}\.xlsx"
                #directory = re.findall(file_path, pattern_dir)
                #print(directory)
                sheet_raw_invoice.cell(row=iterations, column=10).hyperlink = ( "Facturas2022.xlsx#" + months_dict[get_month_pattern] + "!K{}".format(link_sheet))
                print("Facturas2022.xlsx#" + str(months_dict[get_month_pattern]) + "!K{}".format(link_sheet))
                exeption_file = True
            # Checks for the next row
            iterations += 1
        
    # In case the month is incorrect, it will return an error message
    else:
        print("Error, no month {}".format(get_month_pattern[1:3]))
        return None

    # Close and save the document
    close_path(file_path, working_sheet)

